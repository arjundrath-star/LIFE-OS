#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, cast

BASE = Path('/home/Arjun/command-center/Portable Charging')
OUTREACH = BASE / 'Gmail Outreach'
LOG_DIR = OUTREACH / 'send_logs'
LOG_DIR.mkdir(parents=True, exist_ok=True)
RUN_LOG = BASE / '_Run Log.md'
MAIN_CSV = BASE / 'Leads' / 'RVH_Charging_Lead_Pipeline.csv'
ACTIVE_CSV = BASE / 'Leads' / 'Active Leads.csv'
MAIN_XLSX = BASE / 'Leads' / 'RVH_Charging_Lead_Pipeline.xlsx'
ACTIVE_XLSX = BASE / 'Leads' / 'Active Leads.xlsx'
RATH = Path('/home/Arjun/rathworkspace')
GOOGLE_PY = Path('/home/Arjun/.hermes/google-venv/bin/python')
SENDER = 'operator@example.com'

ENV = {
    **os.environ,
    'GOOGLE_WORKSPACE_CLI_CONFIG_DIR': str(Path.home() / '.config/gws-arjun'),
    'PATH': str(Path.home() / '.npm-global/bin') + ':' + str(Path.home() / '.local/bin') + ':/usr/local/bin:/usr/bin:/bin',
}


def now_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00', 'Z')


def today_short() -> str:
    n = datetime.now()
    return f'{n.month}/{n.day}/{str(n.year)[2:]}'


def append_text(existing: Any, note: str) -> str:
    existing = str(existing or '').strip()
    if note in existing:
        return existing
    return f'{existing} {note}'.strip() if existing else note


def event(run_id: str, kind: str, status: str, summary: str, level: str = 'info', artifact: tuple[str, str, str] | None = None) -> None:
    cmd = [
        './node_modules/.bin/tsx', 'scripts/agent-event.ts',
        '--agent', 'portable-charging-outreach-sender',
        '--run', run_id,
        '--kind', kind,
        '--status', status,
        '--summary', summary,
        '--level', level,
        '--trigger-type', 'manual',
        '--trigger-source', 'Hermes Telegram approval',
    ]
    if artifact:
        typ, title, uri = artifact
        cmd += ['--artifact-type', typ, '--artifact-title', title, '--artifact-uri', uri]
    subprocess.run(cmd, cwd=str(RATH), env=ENV, capture_output=True, text=True, timeout=60)


def parse_packet(packet: Path, expected_count: int | None) -> list[dict[str, str]]:
    text = packet.read_text(encoding='utf-8')
    # Headers may use em dashes as Markdown separators. Outgoing fields must not.
    parts = re.split(r'^###\s+(PC-\d{8}-[A-Z]?\d{2})\s+[—-]\s+(.+?)\s*$', text, flags=re.M)
    drafts: list[dict[str, str]] = []
    for i in range(1, len(parts), 3):
        draft_id = parts[i].strip()
        venue = parts[i + 1].strip()
        section = parts[i + 2]
        if '\n---' in section:
            section = section.split('\n---', 1)[0]
        lines = section.strip('\n').splitlines()
        to = subject = None
        body_start = None
        for idx, line in enumerate(lines):
            if line.startswith('To: '):
                to = line[4:].strip()
            elif line.startswith('Subject: '):
                subject = line[9:].strip()
                body_start = idx + 1
                if body_start < len(lines) and lines[body_start] == '':
                    body_start += 1
                break
        if not to or not subject or body_start is None:
            raise SystemExit(f'Could not parse {draft_id}')
        body = '\n'.join(lines[body_start:]).rstrip() + '\n'
        outward = to + '\n' + subject + '\n' + body
        if '—' in outward:
            raise SystemExit(f'ABORT: outgoing em dash found in {draft_id}')
        if '[' in body and ']' in body and 'the operating entity' not in body:
            raise SystemExit(f'ABORT: possible placeholder in {draft_id}')
        drafts.append({'draft_id': draft_id, 'venue': venue, 'to': to, 'subject': subject, 'body': body})
    if expected_count is not None and len(drafts) != expected_count:
        raise SystemExit(f'Expected {expected_count} drafts, parsed {len(drafts)}')
    recipients = [d['to'].lower() for d in drafts]
    dupes = sorted({x for x in recipients if recipients.count(x) > 1})
    if dupes:
        raise SystemExit(f'Duplicate recipients: {dupes}')
    return drafts


def parse_gws_json(out: str) -> dict[str, Any]:
    m = re.search(r'\{.*\}', out, re.S)
    if not m:
        return {}
    try:
        return json.loads(m.group(0))
    except Exception:
        return {}


def run_gws(args: list[str], timeout: int = 180) -> tuple[int, str, dict[str, Any]]:
    r = subprocess.run(['gws', *args], capture_output=True, text=True, env=ENV, cwd=str(BASE), timeout=timeout)
    combined = (r.stdout or '') + (('\nSTDERR:\n' + r.stderr) if r.stderr else '')
    return r.returncode, combined, parse_gws_json(combined)


def send_draft(draft: dict[str, str]) -> dict[str, Any]:
    rc, out, parsed = run_gws(['gmail', '+send', '--to', draft['to'], '--subject', draft['subject'], '--body', draft['body']], timeout=180)
    message_id = parsed.get('id') or parsed.get('message', {}).get('id')
    thread_id = parsed.get('threadId') or parsed.get('message', {}).get('threadId')
    labels = parsed.get('labelIds') or parsed.get('message', {}).get('labelIds') or []
    result: dict[str, Any] = {'returncode': rc, 'message_id': message_id, 'thread_id': thread_id, 'labels': labels, 'raw_output_tail': out[-3000:]}
    if rc != 0 or not message_id:
        result['status'] = 'failed'
        return result
    vrc, vout, _ = run_gws(['gmail', '+read', '--id', message_id, '--format', 'json'], timeout=180)
    result['verify_returncode'] = vrc
    result['verified_sent_label'] = 'SENT' in vout or '"SENT"' in vout or 'SENT' in labels
    result['verify_output_tail'] = vout[-2000:]
    result['status'] = 'sent' if result['verified_sent_label'] else 'sent_unverified'
    return result


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        fieldnames = cast(list[str], reader.fieldnames or [])
        return fieldnames, list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader(); writer.writerows(rows)


def update_csv_files(sent: list[dict[str, Any]], batch: str) -> dict[str, int]:
    sent_by_venue = {d['venue']: d for d in sent if d.get('status') in ('sent', 'sent_unverified')}
    main_fields, main_rows = read_csv(MAIN_CSV)
    active_fields, active_rows = read_csv(ACTIVE_CSV)
    active_by_venue = {r.get('Venue', ''): r for r in active_rows}
    main_updated = active_added = active_updated = 0
    today = today_short()
    for row in main_rows:
        item = sent_by_venue.get(row.get('Venue', ''))
        if not item:
            continue
        note = f"Sent from {SENDER} {today}; approved draft {item['draft_id']}; subject '{item['subject']}'; Gmail message {item.get('message_id')}; thread {item.get('thread_id')}."
        for col, val in {
            'Active Lead': 'Yes',
            'Moved to Active': 'Yes',
            'Active Lead Batch': batch,
            'Active Lead Added At': row.get('Active Lead Added At') or today,
            'Active Outreach Status': f"Venue outreach sent {item['sent_at']} via Gmail; draft {item['draft_id']}; message {item.get('message_id') or ''}",
        }.items():
            if col in row:
                row[col] = val
        if 'Notes' in row:
            row['Notes'] = append_text(row.get('Notes'), note)
        main_updated += 1
        ar = active_by_venue.get(row.get('Venue', ''))
        if ar is None:
            ar = {k: '' for k in active_fields}
            active_rows.append(ar)
            active_by_venue[row.get('Venue', '')] = ar
            active_added += 1
        else:
            active_updated += 1
        mapping = {
            'Active batch': batch,
            'Venue': row.get('Venue', ''),
            'Category': row.get('Category', ''),
            'City': row.get('City', ''),
            'Region': row.get('Region', ''),
            'Address': row.get('Address', ''),
            'Phone': row.get('Phone', ''),
            'Website': row.get('Website', ''),
            'Decision-maker role': row.get('Decision-maker role', ''),
            'Decision-maker name': row.get('Decision-maker name', ''),
            'Email': item.get('to') or row.get('Email', ''),
            'Email confidence': row.get('Email confidence (High/Med/Low)', ''),
            'Hunter verification evidence': row.get('Email Enrichment Notes') or row.get('Email source (Hunter/scraped/inferred)', ''),
            'Fit score': row.get('Fit score (High/Good/Fair/Low)', ''),
            'Tier': row.get('Tier (Pilot/Fast cash/Anchor/Hold)', ''),
            'Captivity rationale': row.get('Captivity rationale', ''),
            'Outreach pattern': 'Email first -> monitor reply/bounce -> follow up or call',
            'Email first status': 'Yes',
            'Email first sent at': today,
            'Last touch method': 'Email',
            'Last touch at': today,
            'Next action': 'Monitor for reply/bounce; follow up if no response after 2 business days.',
            'Owner notes': append_text(ar.get('Owner notes'), note),
            'Date moved to active': ar.get('Date moved to active') or today,
            'Additional Emails': row.get('Additional Emails', ''),
            'Email Deliverability': row.get('Email Deliverability', ''),
            'Email Decision-maker Match': row.get('Email Decision-maker Match', ''),
            'Email Enrichment Notes': row.get('Email Enrichment Notes', ''),
        }
        for k, v in mapping.items():
            if k in ar:
                ar[k] = v
    write_csv(MAIN_CSV, main_fields, main_rows)
    write_csv(ACTIVE_CSV, active_fields, active_rows)
    return {'main_csv_rows': main_updated, 'active_csv_added': active_added, 'active_csv_updated': active_updated, 'active_csv_total': len(active_rows)}


def apply_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> int:
    from openpyxl import load_workbook
    fields, rows = read_csv(csv_path)
    wb = load_workbook(xlsx_path)
    ws = cast(Any, wb.active)
    ws.delete_rows(1, ws.max_row)
    for c, h in enumerate(fields, start=1):
        ws.cell(1, c).value = h
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(fields, start=1):
            ws.cell(r_idx, c_idx).value = row.get(h, '')
    wb.save(xlsx_path)
    return len(rows)


def monitor_after_send() -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, q in {
        'bounces_last_hour': 'from:(mailer-daemon OR postmaster) newer_than:1h',
        'auto_replies_last_hour': 'subject:("Automatic reply" OR "Out of Office" OR "OOO") newer_than:1h',
        'charging_replies_last_hour': 'in:inbox newer_than:1h ("charging amenity" OR "the operating entity" OR "portable charging") -from:mailer-daemon -from:postmaster',
    }.items():
        rc, raw, data = run_gws(['gmail', 'users', 'messages', 'list', '--params', json.dumps({'userId': 'me', 'q': q, 'maxResults': 50})], timeout=180)
        out[key] = {'returncode': rc, 'count': len(data.get('messages', [])) if isinstance(data, dict) else None, 'raw_tail': raw[-1000:] if rc else ''}
    return out


def save_json_atomic(path: Path, data: dict[str, Any]) -> None:
    tmp = path.with_suffix(path.suffix + '.tmp')
    tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    tmp.replace(path)


def write_csv_log(path: Path, sent: list[dict[str, Any]]) -> None:
    with path.open('w', encoding='utf-8', newline='') as f:
        fields = ['draft_id','venue','to','subject','status','sent_at','message_id','thread_id']
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for d in sent:
            writer.writerow({k: d.get(k, '') for k in fields})


def append_run_log(data: dict[str, Any], packet: Path, csv_log: Path) -> None:
    sent = [d for d in data.get('drafts', []) if d.get('status') in ('sent', 'sent_unverified')]
    lines = [
        f"\n## {now_z()} - Approved outreach sender run\n",
        f"- Packet: `{packet}`.",
        f"- Sent {len(sent)} approved drafts from `{SENDER}` with {data.get('cadence_seconds')} seconds between sends.",
        "- Outgoing em dash QA: subjects and bodies scanned before send; zero em dashes found.",
        f"- Log: `{data.get('log_path')}`; CSV: `{csv_log}`.",
        f"- Sheet update: {data.get('sheet_update')}.",
        f"- Drive push return code: {data.get('push', {}).get('returncode')}.",
    ]
    RUN_LOG.write_text(RUN_LOG.read_text(encoding='utf-8').rstrip() + '\n' + '\n'.join(lines) + '\n', encoding='utf-8')


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--packet', required=True)
    ap.add_argument('--expected-count', type=int)
    ap.add_argument('--batch', required=True)
    ap.add_argument('--cadence-seconds', type=int, default=int(os.environ.get('PC_SEND_CADENCE_SECONDS', '180')))
    ap.add_argument('--run-id', default=f'pc-send-{datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")}')
    ap.add_argument('--approval-source', help='Required for live sends: short auditable note naming the user approval/source message.')
    ap.add_argument('--dry-run', action='store_true')
    ns = ap.parse_args()

    if not ns.dry_run and not ns.approval_source:
        raise SystemExit('ABORT: live sends require --approval-source. Draft packets alone are not approval.')

    packet = Path(ns.packet)
    drafts = parse_packet(packet, ns.expected_count)
    log_path = LOG_DIR / f'{ns.run_id}.json'
    csv_log = LOG_DIR / f'{ns.run_id}.csv'

    if ns.dry_run:
        print(json.dumps({'dry_run': True, 'run_id': ns.run_id, 'count': len(drafts), 'cadence_seconds': ns.cadence_seconds, 'packet': str(packet), 'drafts': [{k: v for k, v in d.items() if k != 'body'} for d in drafts]}, indent=2, ensure_ascii=False))
        return 0

    event(ns.run_id, 'started', 'running', f'Approved send run started for {len(drafts)} drafts', 'info', ('review_packet', packet.name, str(packet)))
    data: dict[str, Any] = {'status': 'started', 'started_at': now_z(), 'cadence_seconds': ns.cadence_seconds, 'sender': SENDER, 'packet': str(packet), 'batch': ns.batch, 'run_id': ns.run_id, 'approval_source': ns.approval_source, 'log_path': str(log_path), 'drafts': []}
    if log_path.exists():
        data = json.loads(log_path.read_text(encoding='utf-8'))
    done = {d['draft_id'] for d in data.get('drafts', []) if d.get('status') in ('sent', 'sent_unverified')}

    for idx, draft in enumerate(drafts, start=1):
        if draft['draft_id'] in done:
            print(f'[{now_z()}] skip already sent {draft["draft_id"]}', flush=True)
            continue
        event(ns.run_id, 'sent', 'running', f'Sending {idx}/{len(drafts)} {draft["draft_id"]} to {draft["to"]}')
        record = {k: v for k, v in draft.items() if k != 'body'}
        record['attempted_at'] = now_z()
        print(f'[{record["attempted_at"]}] sending {idx}/{len(drafts)} {draft["draft_id"]} to {draft["to"]}', flush=True)
        record.update(send_draft(draft))
        if record['status'] in ('sent', 'sent_unverified'):
            record['sent_at'] = now_z()
            event(ns.run_id, 'sent', 'running', f'Sent {draft["draft_id"]}; Gmail message {record.get("message_id")}', 'success')
        data.setdefault('drafts', []).append(record)
        save_json_atomic(log_path, data)
        if record['status'] == 'failed':
            data['status'] = 'failed'; data['failed_at'] = now_z(); save_json_atomic(log_path, data)
            event(ns.run_id, 'failed', 'failed', f'Failed on {draft["draft_id"]}; see {log_path}', 'error', ('send_log', log_path.name, str(log_path)))
            print(json.dumps(record, indent=2, ensure_ascii=False), flush=True)
            return 1
        if idx < len(drafts):
            print(f'[{now_z()}] waiting {ns.cadence_seconds}s before next send', flush=True)
            time.sleep(ns.cadence_seconds)

    sent = [d for d in data.get('drafts', []) if d.get('status') in ('sent', 'sent_unverified')]
    event(ns.run_id, 'spreadsheet_update', 'running', 'Updating MAIN and Active Leads after verified sends')
    data['sheet_update'] = update_csv_files(sent, ns.batch)
    data['sheet_update']['main_xlsx_rows'] = apply_csv_to_xlsx(MAIN_CSV, MAIN_XLSX)
    data['sheet_update']['active_xlsx_rows'] = apply_csv_to_xlsx(ACTIVE_CSV, ACTIVE_XLSX)
    fmt = subprocess.run([str(GOOGLE_PY), 'apply_visual_status_formatting.py'], capture_output=True, text=True, cwd=str(BASE), env=ENV, timeout=300)
    push = subprocess.run([str(GOOGLE_PY), 'sync_drive_spreadsheets.py', 'push'], capture_output=True, text=True, cwd=str(BASE), env=ENV, timeout=300)
    data['format'] = {'returncode': fmt.returncode, 'tail': ((fmt.stdout or '') + (fmt.stderr or ''))[-2000:]}
    data['push'] = {'returncode': push.returncode, 'tail': ((push.stdout or '') + (push.stderr or ''))[-3000:]}
    event(ns.run_id, 'bounce_check', 'running', 'Checking fresh bounces and auto replies')
    data['post_send_monitor'] = monitor_after_send()
    data['status'] = 'completed'
    data['completed_at'] = now_z()
    save_json_atomic(log_path, data)
    write_csv_log(csv_log, sent)
    append_run_log(data, packet, csv_log)
    event(ns.run_id, 'completed', 'completed', f'Sent {len(sent)}/{len(drafts)} approved emails; Drive push rc {data["push"]["returncode"]}', 'success', ('send_log', log_path.name, str(log_path)))
    print(json.dumps({'status': data['status'], 'run_id': ns.run_id, 'sent_count': len(sent), 'expected': len(drafts), 'cadence_seconds': ns.cadence_seconds, 'log': str(log_path), 'sheet_update': data['sheet_update'], 'push_returncode': data['push']['returncode'], 'post_send_monitor': data['post_send_monitor']}, indent=2, ensure_ascii=False), flush=True)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
